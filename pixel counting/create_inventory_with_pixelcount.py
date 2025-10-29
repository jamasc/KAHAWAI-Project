# this code explores how to add scores to the sheets

import random
from tqdm import tqdm
from openpyxl import load_workbook
import importlib
import count_pixels
importlib.reload(count_pixels)
from count_pixels import countPixels

def make_path(path):
    path = path.replace('D:', '/mnt/lustre/koa/scratch/jans26/streamflow/images', 1)
    path = path.replace('\\', '/')
    return path

# here goes the loop and getting the image and scores logic
file_paths = [r'/mnt/lustre/koa/scratch/jans26/streamflow/data/image_inventory_cam_1000.xlsx',
              r'/mnt/lustre/koa/scratch/jans26/streamflow/data/image_inventory_cam_1001.xlsx',
              r'/mnt/lustre/koa/scratch/jans26/streamflow/data/image_inventory_cam_1002.xlsx',
              r'/mnt/lustre/koa/scratch/jans26/streamflow/data/image_inventory_cam_1003.xlsx',
              r'/mnt/lustre/koa/scratch/jans26/streamflow/data/image_inventory_cam_1004.xlsx',
              r'/mnt/lustre/koa/scratch/jans26/streamflow/data/image_inventory_cam_1005.xlsx',
              r'/mnt/lustre/koa/scratch/jans26/streamflow/data/image_inventory_cam_1006.xlsx'
             ]
sheet_names = ['Label_1', 'Label_2', 'Label_3', 'Label_4', 'Label_5', 'Label_6']
prompts = ['water', 'rocks', 'river', 'riverbed', 'turbulences', 'stream', 'water rush', 'flowing water', 'glossy', 'dry rocks', 'wet', 'wet rocks', 'water surface', 'streamflow']

def main():
    for file_path in tqdm(file_paths, desc='Camera Files'):
        wb = load_workbook(file_path)
        
        for sheet_name in tqdm(sheet_names, desc='Label Sheets', leave=False):
            ws = wb[sheet_name]
            num_rows = ws.max_row
            # === MAP EXISTING HEADERS ===
            header_map = {}  # column title → column index
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    header_map[header] = col
                    
            for row_index in tqdm(range(2, num_rows + 1), desc='Rows', leave=False):
                image_path = make_path(ws.cell(row=row_index, column=5).value)
                scores = countPixels(prompts, image_path)
                current_max_col = ws.max_column
                for prompt, score in scores.items():
                    if prompt in header_map:
                        col_idx = header_map[prompt]
                    else:
                        current_max_col += 1
                        col_idx = current_max_col
                        ws.cell(row=1, column=col_idx).value = prompt
                        header_map[prompt] = col_idx
                        
                    ws.cell(row=row_index, column=col_idx).value = score
            
            wb.save(file_path)

if __name__ == "__main__":
    main()